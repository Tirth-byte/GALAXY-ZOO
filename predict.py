"""
Galaxy Image Classification - Prediction Pipeline
===================================================
Classify new galaxy images and generate detailed CSV/Excel output.
"""

import os
import sys
from datetime import datetime
from pathlib import Path

import torch
import torch.nn.functional as F
import pandas as pd
from PIL import Image
from tqdm import tqdm

import config
from dataset import get_val_transforms
from model import GalaxyClassifier


# ─── Load Trained Model ─────────────────────────────────────────────────────────

def load_model(model_path=None):
    """
    Load trained model from checkpoint.

    Returns:
        model, device, class_names
    """
    if model_path is None:
        model_path = config.BEST_MODEL_PATH

    if not os.path.exists(model_path):
        print(f"  ❌  Model not found: {model_path}")
        print("  💡  Run `python train.py` first to train the model.")
        sys.exit(1)

    # Select device
    if torch.cuda.is_available():
        device = torch.device("cuda")
    elif hasattr(torch.backends, "mps") and torch.backends.mps.is_available():
        device = torch.device("mps")
    else:
        device = torch.device("cpu")

    # Load checkpoint
    checkpoint = torch.load(model_path, map_location=device, weights_only=False)
    class_names = checkpoint["class_names"]
    num_classes = checkpoint["num_classes"]

    # Build model and load weights
    model = GalaxyClassifier(num_classes=num_classes, pretrained=False)
    model.load_state_dict(checkpoint["model_state_dict"])
    model = model.to(device)
    model.eval()

    val_acc = checkpoint.get("best_val_acc", 0)
    print(f"  ✅  Model loaded (val accuracy: {val_acc:.1%})")
    print(f"  🏷️  Classes: {class_names}")
    print(f"  🖥️  Device: {device}")

    return model, device, class_names


# ─── Classify Single Image ──────────────────────────────────────────────────────

@torch.no_grad()
def classify_image(model, image_path, device, class_names, transform):
    """
    Classify a single galaxy image.

    Returns dict with classification details.
    """
    # Load and transform image
    try:
        image = Image.open(image_path).convert("RGB")
    except Exception as e:
        return {"error": str(e)}

    input_tensor = transform(image).unsqueeze(0).to(device)

    # Forward pass
    logits = model(input_tensor)
    probabilities = F.softmax(logits, dim=1).squeeze().cpu().numpy()

    # Primary prediction
    primary_idx = probabilities.argmax()
    primary_class = class_names[primary_idx]
    confidence = probabilities[primary_idx] * 100

    # Top-3 predictions
    top3_indices = probabilities.argsort()[::-1][:3]
    top3 = [(class_names[i], probabilities[i] * 100) for i in top3_indices]

    # Generate morphology summary
    summary = _generate_summary(primary_class, confidence, top3)

    # Build result dictionary
    result = {
        "image_name": os.path.basename(image_path),
        "image_path": str(image_path),
        "primary_class": primary_class,
        "confidence_pct": round(confidence, 2),
        "top_2_class": top3[1][0] if len(top3) > 1 else "",
        "top_2_confidence": round(top3[1][1], 2) if len(top3) > 1 else 0,
        "top_3_class": top3[2][0] if len(top3) > 2 else "",
        "top_3_confidence": round(top3[2][1], 2) if len(top3) > 2 else 0,
    }

    # Add individual class probabilities
    for i, name in enumerate(class_names):
        col = f"prob_{name.lower().replace(' ', '_').replace('/', '_')}"
        result[col] = round(probabilities[i] * 100, 2)

    result["morphology_summary"] = summary
    result["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    return result


def _generate_summary(primary_class, confidence, top3):
    """Generate a human-readable morphology summary."""
    cls_lower = primary_class.lower()
    conf_str = f"{confidence:.1f}%"

    # Base descriptions
    desc_map = {
        "completely round smooth": f"Elliptical galaxy with a spherical shape ({conf_str} confident). Smooth featureless profile with no visible disk or arms.",
        "in-between smooth": f"Elliptical galaxy, slightly elongated ({conf_str} confident). Smooth profile between round and cigar-shaped.",
        "cigar-shaped smooth": f"Highly elongated elliptical galaxy ({conf_str} confident). Stretched cigar-like morphology.",
        "edge-on disk": f"Disk galaxy viewed edge-on ({conf_str} confident). Narrow band appearance with possible dust lane.",
        "barred spiral": f"Spiral galaxy with a prominent central bar ({conf_str} confident). Bar structure connects spiral arms through the nucleus.",
        "unbarred spiral": f"Spiral galaxy without a central bar ({conf_str} confident). Arms emerge directly from the central bulge.",
        "irregular / merger": f"Irregular or merging galaxy system ({conf_str} confident). Disturbed morphology, possibly interacting or post-merger.",
    }

    summary = desc_map.get(cls_lower, f"{primary_class} galaxy ({conf_str} confident).")

    # If confidence is low, note uncertainty
    if confidence < 50:
        alt = top3[1][0] if len(top3) > 1 else "unknown"
        summary += f" ⚠️ Low confidence — could also be {alt} ({top3[1][1]:.1f}%)."

    return summary


# ─── Scan & Classify All Images ─────────────────────────────────────────────────

def classify_folder(folder_path, model, device, class_names):
    """
    Classify all images in a folder.

    Returns list of result dicts.
    """
    exts = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff", ".webp"}
    image_files = sorted([
        f for f in Path(folder_path).rglob("*")
        if f.suffix.lower() in exts and not f.name.startswith(".")
    ])

    if not image_files:
        print(f"  ⚠️  No images found in: {folder_path}")
        print(f"      Supported formats: {', '.join(exts)}")
        return []

    print(f"  📷  Found {len(image_files)} images to classify\n")

    transform = get_val_transforms()
    results = []

    for img_path in tqdm(image_files, desc="  🔬  Classifying", ncols=80):
        result = classify_image(model, str(img_path), device, class_names, transform)
        if "error" not in result:
            results.append(result)
            # Print inline result
            symbol = _class_symbol(result["primary_class"])
            print(f"    {symbol}  {result['image_name']:40s} → "
                  f"{result['primary_class']} ({result['confidence_pct']:.1f}%)")
        else:
            print(f"    ❌  {img_path.name}: {result['error']}")

    return results


def _class_symbol(class_name):
    """Return emoji for each class."""
    symbols = {
        "completely round smooth": "🟡",
        "in-between smooth": "🟠",
        "cigar-shaped smooth": "🔴",
        "edge-on disk": "📏",
        "barred spiral": "🌀",
        "unbarred spiral": "💫",
        "irregular / merger": "💥",
    }
    return symbols.get(class_name.lower(), "🔵")


# ─── Generate CSV & Excel Output ────────────────────────────────────────────────

def save_results(results, output_dir=None):
    """
    Save classification results to CSV and Excel files.

    Returns paths to saved files.
    """
    if output_dir is None:
        output_dir = config.OUTPUT_DIR
    os.makedirs(output_dir, exist_ok=True)

    if not results:
        print("  ⚠️  No results to save.")
        return None, None

    df = pd.DataFrame(results)

    # Reorder columns for clarity
    priority_cols = [
        "image_name", "primary_class", "confidence_pct",
        "top_2_class", "top_2_confidence",
        "top_3_class", "top_3_confidence",
    ]
    prob_cols = [c for c in df.columns if c.startswith("prob_")]
    other_cols = ["morphology_summary", "timestamp", "image_path"]

    ordered = [c for c in priority_cols + prob_cols + other_cols if c in df.columns]
    remaining = [c for c in df.columns if c not in ordered]
    df = df[ordered + remaining]

    # ── CSV ──
    csv_path = os.path.join(output_dir, "classification_results.csv")
    df.to_csv(csv_path, index=False)
    print(f"\n  📄  CSV saved: {csv_path}")

    # ── Excel with formatting ──
    xlsx_path = os.path.join(output_dir, "classification_results.xlsx")
    _save_formatted_excel(df, xlsx_path)
    print(f"  📊  Excel saved: {xlsx_path}")

    # ── Print summary stats ──
    _print_summary_stats(df)

    return csv_path, xlsx_path


def _save_formatted_excel(df, xlsx_path):
    """Save DataFrame to Excel with formatting and conditional colors."""
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Classifications")

        writer.book
        worksheet = writer.sheets["Classifications"]

        # Import openpyxl styles
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Style constants
        header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        data_font = Font(name="Calibri", size=10)
        border = Border(
            bottom=Side(style="thin", color="D0D0D0"),
            right=Side(style="thin", color="D0D0D0"),
        )

        # Confidence color coding
        high_conf = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                                fill_type="solid")  # Green
        mid_conf = PatternFill(start_color="FFEB9C", end_color="FFEB9C",
                               fill_type="solid")   # Yellow
        low_conf = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Red

        # Format headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Format data rows
        for row_idx in range(2, len(df) + 2):
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = border
                cell.alignment = Alignment(vertical="center")

                # Color-code confidence column
                col_name = df.columns[col_idx - 1]
                if col_name == "confidence_pct":
                    val = cell.value
                    if val is not None:
                        if val >= 80:
                            cell.fill = high_conf
                        elif val >= 50:
                            cell.fill = mid_conf
                        else:
                            cell.fill = low_conf
                    cell.alignment = Alignment(horizontal="center")
                elif col_name.startswith("prob_"):
                    cell.number_format = "0.00"
                    cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column widths
        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = max(
                len(str(col_name)),
                df[col_name].astype(str).str.len().max() if len(df) > 0 else 0
            )
            width = min(max_len + 4, 50)
            worksheet.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header row
        worksheet.freeze_panes = "A2"

        # ── Summary Sheet ──
        summary_df = df.groupby("primary_class").agg(
            count=("image_name", "count"),
            avg_confidence=("confidence_pct", "mean"),
            min_confidence=("confidence_pct", "min"),
            max_confidence=("confidence_pct", "max"),
        ).reset_index()
        summary_df.columns = ["Galaxy Type", "Count",
                              "Avg Confidence (%)", "Min Confidence (%)", "Max Confidence (%)"]
        summary_df = summary_df.sort_values("Count", ascending=False)
        summary_df.to_excel(writer, index=False, sheet_name="Summary")

        # Format summary headers
        ws_summary = writer.sheets["Summary"]
        for col_idx in range(1, len(summary_df.columns) + 1):
            cell = ws_summary.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for col_idx in range(1, len(summary_df.columns) + 1):
            ws_summary.column_dimensions[get_column_letter(col_idx)].width = 22
        ws_summary.freeze_panes = "A2"


def _print_summary_stats(df):
    """Print summary statistics to console."""
    print("\n  📊  Classification Summary:")
    print("  " + "─" * 56)

    counts = df["primary_class"].value_counts()
    total = len(df)

    for cls, count in counts.items():
        avg_conf = df[df["primary_class"] == cls]["confidence_pct"].mean()
        symbol = _class_symbol(cls)
        bar = "█" * int(count / total * 30)
        print(
            f"    {symbol}  {cls:28s} │ {count:4d} ({count/total*100:5.1f}%) avg conf: {avg_conf:5.1f}%  {bar}")

    print("  " + "─" * 56)
    print(f"    Total images: {total}")
    print(f"    Average confidence: {df['confidence_pct'].mean():.1f}%")
    high_conf = (df["confidence_pct"] >= 80).sum()
    print(f"    High confidence (≥80%): {high_conf}/{total} ({high_conf/total*100:.0f}%)")


# ─── Main ────────────────────────────────────────────────────────────────────────

def main():
    print()
    print("╔" + "═" * 58 + "╗")
    print("║    🔭  Galaxy Image Classification — Prediction          ║")
    print("╚" + "═" * 58 + "╝")
    print()

    # Parse optional folder argument
    folder = config.NEW_IMAGES_DIR
    if len(sys.argv) > 1:
        folder = sys.argv[1]
        if not os.path.exists(folder):
            print(f"  ❌  Folder not found: {folder}")
            sys.exit(1)

    print(f"  📂  Image folder: {folder}")

    # Check for images
    if not os.path.exists(folder) or not any(Path(folder).rglob("*")):
        print(f"\n  ⚠️  No images found in '{folder}'")
        print(f"  💡  Place your galaxy images (JPG, PNG, etc.) in:")
        print(f"      {config.NEW_IMAGES_DIR}/")
        print(f"  💡  Or specify a folder: python predict.py /path/to/images")
        return

    # Load model
    print(f"\n  🧠  Loading trained model...")
    model, device, class_names = load_model()

    # Classify all images
    print(f"\n  🔬  Classifying images in: {folder}")
    results = classify_folder(folder, model, device, class_names)

    if not results:
        print("\n  ⚠️  No images were successfully classified.")
        return

    # Save results
    csv_path, xlsx_path = save_results(results)

    print()
    print("  ✅  Classification complete!")
    print(f"  📄  CSV:   {csv_path}")
    print(f"  📊  Excel: {xlsx_path}")
    print(f"\n  💡  Open the Excel file for color-coded results and summary sheet.")
    print()


if __name__ == "__main__":
    main()
