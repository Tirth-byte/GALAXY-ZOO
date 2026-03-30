"""
Galaxy Zoo 1 — Deep Neural Network Classifier
================================================
Train a deep neural network on the GalaxyZoo1_DR_table2.csv tabular data.
Classifies galaxies into 6 morphological categories based on citizen-science
voting probabilities, then outputs full results to CSV + Excel.

Categories:
  0 - Elliptical        (dominant P_EL vote)
  1 - Clockwise Spiral  (dominant P_CW vote)
  2 - Anti-CW Spiral    (dominant P_ACW vote)
  3 - Edge-On            (dominant P_EDGE vote)
  4 - Merger             (dominant P_MG vote)
  5 - Uncertain          (no dominant class / don't know)

Author: Galaxy Zoo Project
"""

import config
from tqdm import tqdm
import seaborn as sns
import matplotlib.pyplot as plt
import os
import sys
import time
import copy
from datetime import datetime

import numpy as np
import pandas as pd
import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import Dataset, DataLoader, WeightedRandomSampler
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import classification_report, confusion_matrix
import matplotlib
matplotlib.use("Agg")


# ═══════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

CSV_PATH = os.path.join(config.BASE_DIR, "GalaxyZoo1_DR_table2.csv")
BATCH_SIZE = 4096
LEARNING_RATE = 3e-4
NUM_EPOCHS = 60
EARLY_STOP_PATIENCE = 10
HIDDEN_DIM = 512
NUM_LAYERS = 4
DROPOUT = 0.3

CLASS_NAMES = [
    "Elliptical",
    "Clockwise Spiral",
    "Anti-CW Spiral",
    "Edge-On",
    "Merger",
    "Uncertain",
]
NUM_CLASSES = len(CLASS_NAMES)

# Feature columns from the CSV
PROB_FEATURES = ["P_EL", "P_CW", "P_ACW", "P_EDGE", "P_DK", "P_MG", "P_CS"]
META_FEATURES = ["NVOTE", "P_EL_DEBIASED", "P_CS_DEBIASED"]


# ═══════════════════════════════════════════════════════════════════════════════
#  DATA LOADING & FEATURE ENGINEERING
# ═══════════════════════════════════════════════════════════════════════════════

def load_and_prepare_data(csv_path):
    """
    Load Galaxy Zoo CSV, engineer features, assign class labels.
    Returns DataFrame with features and labels.
    """
    print("  📂  Loading CSV...")
    df = pd.read_csv(csv_path)
    print(f"      Loaded {len(df):,} galaxies with {len(df.columns)} columns")

    # ── Feature Engineering ──
    print("  🔧  Engineering features...")

    # Core probability features (already in CSV)
    # P_EL, P_CW, P_ACW, P_EDGE, P_DK, P_MG, P_CS

    # Derived features
    df["P_SPIRAL_TOTAL"] = df["P_CW"] + df["P_ACW"]         # Total spiral probability
    df["SPIRAL_ASYMMETRY"] = (df["P_CW"] - df["P_ACW"]).abs()  # CW vs ACW difference
    df["EL_CS_RATIO"] = df["P_EL"] / (df["P_CS"] + 1e-8)    # Elliptical-to-Spiral ratio
    df["CERTAINTY"] = 1.0 - df["P_DK"]                       # How certain voters were
    df["MAX_PROB"] = df[["P_EL", "P_CW", "P_ACW", "P_EDGE", "P_MG"]].max(axis=1)
    df["ENTROPY"] = -sum(
        df[col] * np.log(df[col] + 1e-10) for col in PROB_FEATURES
    )  # Voting entropy (higher = more uncertain)
    df["NVOTE_NORM"] = df["NVOTE"] / df["NVOTE"].max()       # Normalized vote count

    # ── Assign Class Labels ──
    # Use hard assignment based on highest morphology probability
    # with a threshold to separate "Uncertain" cases
    print("  🏷️  Assigning class labels...")

    morph_probs = df[["P_EL", "P_CW", "P_ACW", "P_EDGE", "P_MG"]].values
    max_indices = morph_probs.argmax(axis=1)
    max_values = morph_probs.max(axis=1)

    # Map: 0=Elliptical, 1=CW Spiral, 2=ACW Spiral, 3=Edge-On, 4=Merger
    labels = max_indices.copy()

    # If the max probability is too low (< 0.3) AND P_DK is significant,
    # classify as "Uncertain" (class 5)
    uncertain_mask = (max_values < 0.3) | (df["P_DK"].values > 0.4)
    labels[uncertain_mask] = 5

    df["label"] = labels

    # Print distribution
    print("\n  📊  Class Distribution:")
    print("  " + "─" * 55)
    for i, name in enumerate(CLASS_NAMES):
        count = (df["label"] == i).sum()
        pct = count / len(df) * 100
        bar = "█" * int(pct / 2)
        print(f"    {i} {name:20s} │ {count:7,d} ({pct:5.1f}%) {bar}")
    print("  " + "─" * 55)
    print(f"    Total: {len(df):,}")

    return df


def get_feature_columns():
    """Return the list of all feature column names."""
    return PROB_FEATURES + META_FEATURES + [
        "P_SPIRAL_TOTAL", "SPIRAL_ASYMMETRY", "EL_CS_RATIO",
        "CERTAINTY", "MAX_PROB", "ENTROPY", "NVOTE_NORM",
    ]


# ═══════════════════════════════════════════════════════════════════════════════
#  PYTORCH DATASET
# ═══════════════════════════════════════════════════════════════════════════════

class GalaxyTabularDataset(Dataset):
    """PyTorch Dataset for tabular galaxy features."""

    def __init__(self, features, labels):
        self.features = torch.FloatTensor(features)
        self.labels = torch.LongTensor(labels)

    def __len__(self):
        return len(self.labels)

    def __getitem__(self, idx):
        return self.features[idx], self.labels[idx]


# ═══════════════════════════════════════════════════════════════════════════════
#  DEEP NEURAL NETWORK MODEL
# ═══════════════════════════════════════════════════════════════════════════════

class ResidualBlock(nn.Module):
    """Residual block with BatchNorm, GELU, and Dropout."""

    def __init__(self, dim, dropout=0.3):
        super().__init__()
        self.block = nn.Sequential(
            nn.Linear(dim, dim),
            nn.BatchNorm1d(dim),
            nn.GELU(),
            nn.Dropout(dropout),
            nn.Linear(dim, dim),
            nn.BatchNorm1d(dim),
        )
        self.activation = nn.GELU()
        self.dropout = nn.Dropout(dropout * 0.5)

    def forward(self, x):
        return self.dropout(self.activation(x + self.block(x)))


class GalaxyDNN(nn.Module):
    """
    Deep Neural Network for galaxy morphology classification.

    Architecture:
        Input → FC(in, 512) → BatchNorm → GELU
        → [ResidualBlock(512)] × 4
        → FC(512, 256) → BatchNorm → GELU → Dropout
        → FC(256, num_classes)
    """

    def __init__(self, input_dim, num_classes=NUM_CLASSES, hidden_dim=HIDDEN_DIM,
                 num_layers=NUM_LAYERS, dropout=DROPOUT):
        super().__init__()

        # Input projection
        self.input_layer = nn.Sequential(
            nn.Linear(input_dim, hidden_dim),
            nn.BatchNorm1d(hidden_dim),
            nn.GELU(),
            nn.Dropout(dropout),
        )

        # Residual blocks
        self.res_blocks = nn.ModuleList([
            ResidualBlock(hidden_dim, dropout) for _ in range(num_layers)
        ])

        # Output head
        self.output_head = nn.Sequential(
            nn.Linear(hidden_dim, hidden_dim // 2),
            nn.BatchNorm1d(hidden_dim // 2),
            nn.GELU(),
            nn.Dropout(dropout * 0.5),
            nn.Linear(hidden_dim // 2, num_classes),
        )

    def forward(self, x):
        x = self.input_layer(x)
        for block in self.res_blocks:
            x = block(x)
        return self.output_head(x)


# ═══════════════════════════════════════════════════════════════════════════════
#  TRAINING FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def train_one_epoch(model, loader, criterion, optimizer, device):
    """Train for one epoch."""
    model.train()
    total_loss = 0.0
    correct = 0
    total = 0

    for features, labels in loader:
        features = features.to(device, non_blocking=True)
        labels = labels.to(device, non_blocking=True)

        optimizer.zero_grad()
        outputs = model(features)
        loss = criterion(outputs, labels)
        loss.backward()
        torch.nn.utils.clip_grad_norm_(model.parameters(), max_norm=1.0)
        optimizer.step()

        total_loss += loss.item() * features.size(0)
        _, predicted = outputs.max(1)
        total += labels.size(0)
        correct += predicted.eq(labels).sum().item()

    return total_loss / total, correct / total


@torch.no_grad()
def validate(model, loader, criterion, device):
    """Validate model."""
    model.eval()
    total_loss = 0.0
    correct = 0
    total = 0
    all_preds = []
    all_labels = []

    for features, labels in loader:
        features = features.to(device, non_blocking=True)
        labels = labels.to(device, non_blocking=True)

        outputs = model(features)
        loss = criterion(outputs, labels)

        total_loss += loss.item() * features.size(0)
        _, predicted = outputs.max(1)
        total += labels.size(0)
        correct += predicted.eq(labels).sum().item()

        all_preds.extend(predicted.cpu().numpy())
        all_labels.extend(labels.cpu().numpy())

    return total_loss / total, correct / total, np.array(all_preds), np.array(all_labels)


# ═══════════════════════════════════════════════════════════════════════════════
#  PLOTTING
# ═══════════════════════════════════════════════════════════════════════════════

def plot_confusion_matrix(y_true, y_pred, class_names, save_path):
    """Plot confusion matrix (counts + normalized)."""
    cm = confusion_matrix(y_true, y_pred)
    cm_norm = cm.astype("float") / cm.sum(axis=1, keepdims=True)

    fig, axes = plt.subplots(1, 2, figsize=(20, 8))

    sns.heatmap(cm, annot=True, fmt=",d", cmap="Blues", xticklabels=class_names,
                yticklabels=class_names, ax=axes[0], cbar_kws={"shrink": 0.8})
    axes[0].set_title("Confusion Matrix (Counts)", fontsize=14, fontweight="bold")
    axes[0].set_xlabel("Predicted", fontsize=12)
    axes[0].set_ylabel("Actual", fontsize=12)
    axes[0].tick_params(axis='x', rotation=45)

    sns.heatmap(cm_norm, annot=True, fmt=".3f", cmap="YlOrRd", xticklabels=class_names,
                yticklabels=class_names, ax=axes[1], cbar_kws={"shrink": 0.8})
    axes[1].set_title("Confusion Matrix (Normalized)", fontsize=14, fontweight="bold")
    axes[1].set_xlabel("Predicted", fontsize=12)
    axes[1].set_ylabel("Actual", fontsize=12)
    axes[1].tick_params(axis='x', rotation=45)

    plt.tight_layout()
    plt.savefig(save_path, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  📊  Confusion matrix saved: {save_path}")


def plot_training_history(history, save_path):
    """Plot loss and accuracy curves."""
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))
    epochs = range(1, len(history["train_loss"]) + 1)

    ax1.plot(epochs, history["train_loss"], "b-", linewidth=1.5, label="Train")
    ax1.plot(epochs, history["val_loss"], "r-", linewidth=1.5, label="Val")
    ax1.set_title("Loss", fontsize=14, fontweight="bold")
    ax1.set_xlabel("Epoch")
    ax1.set_ylabel("Loss")
    ax1.legend()
    ax1.grid(True, alpha=0.3)

    ax2.plot(epochs, history["train_acc"], "b-", linewidth=1.5, label="Train")
    ax2.plot(epochs, history["val_acc"], "r-", linewidth=1.5, label="Val")
    ax2.set_title("Accuracy", fontsize=14, fontweight="bold")
    ax2.set_xlabel("Epoch")
    ax2.set_ylabel("Accuracy")
    ax2.legend()
    ax2.grid(True, alpha=0.3)

    plt.tight_layout()
    plt.savefig(save_path, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  📈  Training history saved: {save_path}")


# ═══════════════════════════════════════════════════════════════════════════════
#  FULL-DATASET PREDICTION → CSV + EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

@torch.no_grad()
def predict_all(model, df, feature_cols, scaler, device):
    """
    Run prediction on the ENTIRE dataset and return enriched DataFrame.
    """
    print("\n  🔬  Classifying all 667,944 galaxies...")
    model.eval()

    features = scaler.transform(df[feature_cols].values)
    dataset = GalaxyTabularDataset(features, np.zeros(len(df), dtype=int))
    loader = DataLoader(dataset, batch_size=8192, shuffle=False, num_workers=0)

    all_probs = []
    for batch_features, _ in tqdm(loader, desc="  Predicting", ncols=80):
        batch_features = batch_features.to(device)
        logits = model(batch_features)
        probs = torch.softmax(logits, dim=1).cpu().numpy()
        all_probs.append(probs)

    all_probs = np.concatenate(all_probs, axis=0)

    # Build results DataFrame
    results = pd.DataFrame()
    results["OBJID"] = df["OBJID"].values
    results["RA"] = df["RA"].values
    results["DEC"] = df["DEC"].values
    results["NVOTE"] = df["NVOTE"].values

    # Original probabilities
    for col in PROB_FEATURES:
        results[f"original_{col}"] = df[col].values

    # Predicted class
    pred_indices = all_probs.argmax(axis=1)
    results["predicted_class"] = [CLASS_NAMES[i] for i in pred_indices]
    results["confidence_pct"] = np.round(all_probs.max(axis=1) * 100, 2)

    # Per-class DNN probabilities
    for i, name in enumerate(CLASS_NAMES):
        col = f"DNN_prob_{name.lower().replace(' ', '_').replace('-', '')}"
        results[col] = np.round(all_probs[:, i] * 100, 2)

    # Top-2 and Top-3
    top3 = all_probs.argsort(axis=1)[:, ::-1][:, :3]
    results["top_2_class"] = [CLASS_NAMES[i] for i in top3[:, 1]]
    results["top_2_confidence"] = np.round(
        all_probs[np.arange(len(all_probs)), top3[:, 1]] * 100, 2)
    results["top_3_class"] = [CLASS_NAMES[i] for i in top3[:, 2]]
    results["top_3_confidence"] = np.round(
        all_probs[np.arange(len(all_probs)), top3[:, 2]] * 100, 2)

    # Original Galaxy Zoo flags
    results["GZ1_SPIRAL"] = df["SPIRAL"].values
    results["GZ1_ELLIPTICAL"] = df["ELLIPTICAL"].values
    results["GZ1_UNCERTAIN"] = df["UNCERTAIN"].values

    # Agreement with original GZ1 classification
    gz1_class = np.where(df["SPIRAL"].values == 1, "Spiral",
                         np.where(df["ELLIPTICAL"].values == 1, "Elliptical", "Uncertain"))
    dnn_is_spiral = np.isin(pred_indices, [1, 2, 3])  # CW, ACW, Edge-On are spiral-type
    dnn_is_elliptical = pred_indices == 0
    dnn_simplified = np.where(dnn_is_spiral, "Spiral",
                              np.where(dnn_is_elliptical, "Elliptical", "Uncertain"))
    results["agrees_with_GZ1"] = (gz1_class == dnn_simplified)

    # Morphology summary
    summaries = []
    for i in range(len(results)):
        cls = results.iloc[i]["predicted_class"]
        conf = results.iloc[i]["confidence_pct"]
        desc_map = {
            "Elliptical": f"Elliptical galaxy — smooth, rounded shape ({conf:.1f}% confident)",
            "Clockwise Spiral": f"Spiral galaxy with clockwise arm rotation ({conf:.1f}% confident)",
            "Anti-CW Spiral": f"Spiral galaxy with anti-clockwise arm rotation ({conf:.1f}% confident)",
            "Edge-On": f"Disk galaxy viewed edge-on as a narrow band ({conf:.1f}% confident)",
            "Merger": f"Merging or disturbed galaxy system ({conf:.1f}% confident)",
            "Uncertain": f"Uncertain morphology — no dominant feature ({conf:.1f}% confident)",
        }
        summaries.append(desc_map.get(cls, f"{cls} ({conf:.1f}%)"))
    results["morphology_summary"] = summaries
    results["classification_timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    return results


def save_results(results_df):
    """Save results to CSV and formatted Excel."""
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)

    # ── CSV ──
    csv_path = os.path.join(config.OUTPUT_DIR, "classification_results.csv")
    results_df.to_csv(csv_path, index=False)
    print(f"\n  📄  CSV saved: {csv_path} ({len(results_df):,} rows)")

    # ── Excel ──
    xlsx_path = os.path.join(config.OUTPUT_DIR, "classification_results.xlsx")
    print(f"  📊  Writing Excel file (this may take a minute for 667K rows)...")

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        # Main sheet — write first 100,000 rows (Excel has a ~1M row limit, and
        # formatting 667K rows would be extremely slow)
        sample_size = min(100_000, len(results_df))
        results_df.head(sample_size).to_excel(writer, index=False, sheet_name="Classifications")

        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        ws = writer.sheets["Classifications"]

        # Style headers
        header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        for col_idx in range(1, len(results_df.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Color-code confidence column
        conf_col = list(results_df.columns).index("confidence_pct") + 1
        high = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        mid = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        low = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for row_idx in range(2, min(sample_size + 2, 100002)):
            cell = ws.cell(row=row_idx, column=conf_col)
            if cell.value is not None:
                if cell.value >= 80:
                    cell.fill = high
                elif cell.value >= 50:
                    cell.fill = mid
                else:
                    cell.fill = low

        # Auto-width (approximate)
        for col_idx, col_name in enumerate(results_df.columns, 1):
            width = min(max(len(str(col_name)) + 2, 12), 35)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = "A2"

        # ── Summary Sheet ──
        summary = results_df.groupby("predicted_class").agg(
            count=("OBJID", "count"),
            avg_confidence=("confidence_pct", "mean"),
            min_confidence=("confidence_pct", "min"),
            max_confidence=("confidence_pct", "max"),
            avg_votes=("NVOTE", "mean"),
        ).reset_index()
        summary.columns = ["Galaxy Type", "Count", "Avg Confidence (%)",
                           "Min Confidence (%)", "Max Confidence (%)", "Avg Votes"]
        summary = summary.sort_values("Count", ascending=False)
        summary.to_excel(writer, index=False, sheet_name="Summary")

        ws2 = writer.sheets["Summary"]
        for col_idx in range(1, len(summary.columns) + 1):
            cell = ws2.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws2.column_dimensions[get_column_letter(col_idx)].width = 22
        ws2.freeze_panes = "A2"

        # ── Agreement Sheet ──
        agree_pct = results_df["agrees_with_GZ1"].mean() * 100
        agree_df = pd.DataFrame({
            "Metric": [
                "Total Galaxies",
                "Agrees with Galaxy Zoo 1",
                "Disagrees with Galaxy Zoo 1",
                "Agreement Rate (%)",
            ],
            "Value": [
                f"{len(results_df):,}",
                f"{results_df['agrees_with_GZ1'].sum():,}",
                f"{(~results_df['agrees_with_GZ1']).sum():,}",
                f"{agree_pct:.2f}%",
            ]
        })
        agree_df.to_excel(writer, index=False, sheet_name="GZ1 Agreement")
        ws3 = writer.sheets["GZ1 Agreement"]
        for col_idx in range(1, 3):
            cell = ws3.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            ws3.column_dimensions[get_column_letter(col_idx)].width = 30

    print(f"  📊  Excel saved: {xlsx_path}")
    if sample_size < len(results_df):
        print(f"      ⚠️  Excel contains first {sample_size:,} rows (full data in CSV)")

    # ── Print summary ──
    print("\n  📊  Final Classification Summary:")
    print("  " + "─" * 60)
    for cls in CLASS_NAMES:
        subset = results_df[results_df["predicted_class"] == cls]
        count = len(subset)
        pct = count / len(results_df) * 100
        avg_conf = subset["confidence_pct"].mean() if count > 0 else 0
        bar = "█" * int(pct / 2)
        print(f"    {cls:20s} │ {count:7,d} ({pct:5.1f}%) avg_conf: {avg_conf:5.1f}%  {bar}")
    print("  " + "─" * 60)
    agree = results_df["agrees_with_GZ1"].mean() * 100
    print(f"    Agreement with Galaxy Zoo 1: {agree:.2f}%")


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN TRAINING PIPELINE
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    print()
    print("╔" + "═" * 62 + "╗")
    print("║   🔭  Galaxy Zoo 1 — Deep Neural Network Classification    ║")
    print("║   📄  Dataset: GalaxyZoo1_DR_table2.csv (667,944 galaxies) ║")
    print("╚" + "═" * 62 + "╝")
    print()

    # ── Check CSV ──
    if not os.path.exists(CSV_PATH):
        print(f"  ❌  CSV not found: {CSV_PATH}")
        sys.exit(1)

    # ── Device ──
    # Use CPU for tabular DNN — fast enough and avoids MPS malloc issues
    if torch.cuda.is_available():
        device = torch.device("cuda")
        print(f"  🖥️  GPU: {torch.cuda.get_device_name(0)}")
    else:
        device = torch.device("cpu")
        print("  💻  Using CPU (optimal for tabular DNN)")

    # ══════════════════════════════════════════════════════════════════════
    #  STEP 1: LOAD & PREPARE DATA
    # ══════════════════════════════════════════════════════════════════════
    print("\n  ━━━ Step 1/5: Loading & Preparing Data ━━━")
    df = load_and_prepare_data(CSV_PATH)
    feature_cols = get_feature_columns()
    print(f"\n  📐  Features: {len(feature_cols)} columns")
    print(f"      {feature_cols}")

    X = df[feature_cols].values.astype(np.float32)
    y = df["label"].values

    # Handle any NaN/Inf
    X = np.nan_to_num(X, nan=0.0, posinf=1.0, neginf=0.0)

    # Scale features
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)

    # Train/val split (stratified)
    X_train, X_val, y_train, y_val = train_test_split(
        X_scaled, y, test_size=0.15, stratify=y, random_state=42
    )
    print(f"\n  🔀  Split: {len(X_train):,} train / {len(X_val):,} validation")

    # ══════════════════════════════════════════════════════════════════════
    #  STEP 2: DATA LOADERS
    # ══════════════════════════════════════════════════════════════════════
    print("\n  ━━━ Step 2/5: Creating Data Loaders ━━━")

    train_dataset = GalaxyTabularDataset(X_train, y_train)
    val_dataset = GalaxyTabularDataset(X_val, y_val)

    # Weighted sampler for class imbalance
    class_counts = np.bincount(y_train, minlength=NUM_CLASSES)
    class_weights = 1.0 / (class_counts + 1)
    sample_weights = class_weights[y_train]
    sampler = WeightedRandomSampler(
        sample_weights, num_samples=len(sample_weights), replacement=True)

    train_loader = DataLoader(train_dataset, batch_size=BATCH_SIZE, sampler=sampler,
                              num_workers=0, pin_memory=False, drop_last=True)
    val_loader = DataLoader(val_dataset, batch_size=BATCH_SIZE * 2, shuffle=False,
                            num_workers=0, pin_memory=False)

    print(f"  📦  Batch size: {BATCH_SIZE}")
    print(f"  📦  Train batches: {len(train_loader)} | Val batches: {len(val_loader)}")

    # ══════════════════════════════════════════════════════════════════════
    #  STEP 3: BUILD MODEL
    # ══════════════════════════════════════════════════════════════════════
    print("\n  ━━━ Step 3/5: Building DNN Model ━━━")

    model = GalaxyDNN(
        input_dim=len(feature_cols),
        num_classes=NUM_CLASSES,
        hidden_dim=HIDDEN_DIM,
        num_layers=NUM_LAYERS,
        dropout=DROPOUT,
    ).to(device)

    total_params = sum(p.numel() for p in model.parameters())
    print(f"  🧠  Model: {NUM_LAYERS}-layer Residual DNN")
    print(f"      Input dim: {len(feature_cols)} → Hidden: {HIDDEN_DIM} → Output: {NUM_CLASSES}")
    print(f"      Parameters: {total_params:,}")

    # Loss with class weights
    weight_tensor = torch.FloatTensor(class_weights / class_weights.sum() * NUM_CLASSES).to(device)
    criterion = nn.CrossEntropyLoss(weight=weight_tensor)

    optimizer = optim.AdamW(model.parameters(), lr=LEARNING_RATE, weight_decay=1e-4)
    scheduler = optim.lr_scheduler.CosineAnnealingWarmRestarts(optimizer, T_0=10, T_mult=2)

    print(f"  ⚙️  Optimizer: AdamW (lr={LEARNING_RATE})")
    print(f"  ⚙️  Scheduler: CosineAnnealingWarmRestarts")
    print(f"  ⚙️  Epochs: {NUM_EPOCHS} (early stop patience: {EARLY_STOP_PATIENCE})")

    # ══════════════════════════════════════════════════════════════════════
    #  STEP 4: TRAINING LOOP
    # ══════════════════════════════════════════════════════════════════════
    print("\n  ━━━ Step 4/5: Training ━━━")
    print("  " + "─" * 70)
    print(f"  {'Epoch':>6} │ {'Train Loss':>10} │ {'Val Loss':>10} │ {'Train Acc':>9} │ {'Val Acc':>9} │ {'LR':>9} │")
    print("  " + "─" * 70)

    history = {"train_loss": [], "train_acc": [], "val_loss": [], "val_acc": []}
    best_val_acc = 0.0
    best_model_state = None
    patience_counter = 0
    start_time = time.time()

    for epoch in range(1, NUM_EPOCHS + 1):
        train_loss, train_acc = train_one_epoch(model, train_loader, criterion, optimizer, device)
        val_loss, val_acc, val_preds, val_labels = validate(model, val_loader, criterion, device)
        scheduler.step()

        history["train_loss"].append(train_loss)
        history["train_acc"].append(train_acc)
        history["val_loss"].append(val_loss)
        history["val_acc"].append(val_acc)

        lr = optimizer.param_groups[0]["lr"]
        flag = ""
        if val_acc > best_val_acc:
            best_val_acc = val_acc
            best_model_state = copy.deepcopy(model.state_dict())
            patience_counter = 0
            flag = " ★"
        else:
            patience_counter += 1

        print(f"  {epoch:5d} │ {train_loss:10.4f} │ {val_loss:10.4f} │ {train_acc:8.4f} │ {val_acc:8.4f} │ {lr:9.2e} │{flag}")

        if patience_counter >= EARLY_STOP_PATIENCE:
            print(f"\n  ⏹️  Early stopping at epoch {epoch}")
            break

    total_time = time.time() - start_time
    print("  " + "─" * 70)
    print(f"  ⏱️  Training time: {total_time / 60:.1f} minutes")
    print(f"  🏆  Best validation accuracy: {best_val_acc:.4f} ({best_val_acc * 100:.2f}%)")

    # Save model
    os.makedirs(config.MODEL_DIR, exist_ok=True)
    model_path = os.path.join(config.MODEL_DIR, "galaxy_dnn_best.pth")
    torch.save({
        "model_state_dict": best_model_state,
        "class_names": CLASS_NAMES,
        "num_classes": NUM_CLASSES,
        "feature_cols": feature_cols,
        "input_dim": len(feature_cols),
        "hidden_dim": HIDDEN_DIM,
        "num_layers": NUM_LAYERS,
        "best_val_acc": best_val_acc,
        "scaler_mean": scaler.mean_.tolist(),
        "scaler_scale": scaler.scale_.tolist(),
    }, model_path)
    print(f"  💾  Model saved: {model_path}")

    # Load best model
    model.load_state_dict(best_model_state)

    # Final validation report
    _, _, final_preds, final_labels = validate(model, val_loader, criterion, device)
    print("\n  📋  Classification Report (Validation Set):")
    print("  " + "─" * 60)
    report = classification_report(final_labels, final_preds, target_names=CLASS_NAMES, digits=4)
    for line in report.strip().split("\n"):
        print(f"    {line}")

    # Plots
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)
    plot_confusion_matrix(final_labels, final_preds, CLASS_NAMES, config.CONFUSION_MATRIX_PATH)
    plot_training_history(history, config.TRAINING_HISTORY_PATH)

    # ══════════════════════════════════════════════════════════════════════
    #  STEP 5: PREDICT ALL & SAVE CSV/EXCEL
    # ══════════════════════════════════════════════════════════════════════
    print("\n  ━━━ Step 5/5: Classifying All 667,944 Galaxies ━━━")
    results_df = predict_all(model, df, feature_cols, scaler, device)
    save_results(results_df)

    print()
    print("  " + "═" * 62)
    print("  ✅  COMPLETE! All 667,944 galaxies classified.")
    print(f"  📄  CSV:   {os.path.join(config.OUTPUT_DIR, 'classification_results.csv')}")
    print(f"  📊  Excel: {os.path.join(config.OUTPUT_DIR, 'classification_results.xlsx')}")
    print(f"  📊  Confusion matrix: {config.CONFUSION_MATRIX_PATH}")
    print(f"  📈  Training curves:  {config.TRAINING_HISTORY_PATH}")
    print("  " + "═" * 62)
    print()


if __name__ == "__main__":
    main()
